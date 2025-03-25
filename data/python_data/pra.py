import pandas as pd
import numpy as np
import openpyxl

# --- Data Generation ---

num_rows = 100

# Define headers and subheaders (more complex structure)
headers = [
    "Product", "Product", "Product",
    "Inventory", "Inventory",
    "Sales", "Sales",
    "Supplier", "Supplier"
]
subheaders = [
    "ID", "Name", "Category",
    "Quantity", "Location",
    "Units Sold", "Revenue",
    "Name", "Contact"
]

# Check for length consistency
if len(headers) != len(subheaders):
    raise ValueError("Headers and subheaders lists must have the same length.")


# Category-specific subheaders (THIS IS THE KEY PART)
category_subheaders = {
    "Electronics": ["Brand", "Model", "Warranty (months)"],
    "Clothing": ["Size", "Color", "Material"],
    "Books": ["Author", "Genre", "ISBN"],
    "Food": ["Expiry Date", "Origin", "Weight (g)"],
}

# Generate data for each column
data = {
    ("Product", "ID"): [f"P{i:03}" for i in range(1, num_rows + 1)],  # Product IDs
    ("Product", "Name"): [f"Product {i}" for i in range(1, num_rows + 1)],  # Product Names
    # ("Product", "Category"): assigned later, depends on category subheaders
    ("Inventory", "Quantity"): np.random.randint(10, 200, size=num_rows),
    ("Inventory", "Location"): np.random.choice(["Warehouse A", "Warehouse B", "Store Front"], size=num_rows),
    ("Sales", "Units Sold"): np.random.randint(1, 50, size=num_rows),
    ("Sales", "Revenue"): np.random.randint(100, 5000, size=num_rows),
    ("Supplier", "Name"): [f"Supplier {i}" for i in range(1, num_rows + 1)],  # Supplier names
    ("Supplier", "Contact"): [f"contact{i}@supplier.com" for i in range(1, num_rows + 1)],  # Contact emails
}

# --- Dynamically Generate Category and Sub-Category Columns ---

categories = list(category_subheaders.keys())
data[("Product", "Category")] = np.random.choice(categories, size=num_rows)

# Create a list to hold the expanded data
expanded_data = []

# Iterate through the generated data and add the appropriate category subheaders
for i in range(num_rows):
    row_data = {}
    for (header, subheader), values in data.items():
        row_data[(header, subheader)] = values[i]  # Get the value for this row

    category = row_data[("Product", "Category")]

    for j, sub_cat_header in enumerate(category_subheaders[category]):
        # Generate some sample sub-category data (you can customize this)
        if category == "Electronics":
            if j == 0:  # Brand
                sub_cat_value = np.random.choice(["Samsung", "Apple", "LG"])
            elif j == 1: # Model
                sub_cat_value = f"Model{np.random.randint(100,999)}"
            else:  # Warranty
                sub_cat_value = np.random.randint(6, 25)
        elif category == "Clothing":
            if j == 0: # Size
                sub_cat_value = np.random.choice(["S", "M", "L", "XL"])
            elif j == 1: # Color
                sub_cat_value = np.random.choice(["Red", "Blue", "Green", "Black"])
            else:  # Material
                sub_cat_value = np.random.choice(["Cotton", "Polyester", "Wool"])
        elif category == "Books":
            if j==0: #Author
                sub_cat_value = f"Author {np.random.randint(1,21)}"
            elif j == 1: #Genre
                sub_cat_value = np.random.choice(["Fiction", "Non-Fiction", "Sci-Fi", "Mystery"])
            else: #ISBN
                sub_cat_value = f"ISBN-{np.random.randint(1000000000, 9999999999)}"
        elif category == "Food":
            if j == 0: # Expiry
                sub_cat_value = f"2024-{np.random.randint(1,13):02}-{np.random.randint(1,29):02}"  # Simplified date format
            elif j==1: # Origin
                sub_cat_value = np.random.choice(["USA", "Canada", "Mexico"])
            else: #weight
                sub_cat_value = np.random.randint(100,1001)

        row_data[("Product", category, sub_cat_header)] = sub_cat_value
    expanded_data.append(row_data)


# --- Create DataFrame from Expanded Data ---
# Build a comprehensive list of all column tuples
all_columns = []
for row in expanded_data:
    all_columns.extend(row.keys())
all_columns = list(set(all_columns)) # Remove duplicates

# Create MultiIndex
multi_index = pd.MultiIndex.from_tuples(all_columns)

final_df = pd.DataFrame(expanded_data, columns=multi_index)

# --- Save to XLSX ---
file_path = "inventory_data.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
col_idx = 1
for header in headers:
    sheet.cell(row=1, column=col_idx, value=header)
    # Check for subheaders and span as needed
    if header == "Product":  # Example: Span Product header
        num_subheaders = 3  # "ID", "Name", "Category"
        for cat in categories:
            num_subheaders += len(category_subheaders.get(cat, []))
            
        sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + num_subheaders -1)  # Merge across subheaders
        col_idx += num_subheaders
    elif header == "Inventory" or header == "Sales" or header == "Supplier":
        sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column= col_idx + 1)
        col_idx += 2
    else:
        col_idx +=1

# Write subheaders
col_idx = 1
for header, subheader in zip (headers, subheaders):
    sheet.cell(row=2, column=col_idx, value=subheader)
    col_idx += 1
    
#Write subheader for category    
for category in categories:
        for sub_cat_header in category_subheaders[category]:
            sheet.cell(row=2, column=col_idx, value=sub_cat_header)
            col_idx +=1



# Write data, starting from row 3
for r_idx, row in final_df.iterrows():
    for c_idx, value in enumerate(row, 1):  # Start column index at 1 for openpyxl
        sheet.cell(row=r_idx + 3, column=c_idx, value=value)
        

workbook.save(file_path)
print(f"Data saved to '{file_path}'")
print(final_df.head())
